import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import datetime
from typing import List, Optional


class ExcelManager:
    def __init__(self, filename: str = "test_results.xlsx"):
        self.filename = filename
        self.wb = None
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        if os.path.exists(self.filename):
            self.wb = openpyxl.load_workbook(self.filename)
        else:
            self.wb = openpyxl.Workbook()
            default_sheet = self.wb.active
            if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                self.wb.remove(default_sheet)

    def save_results(self, results: List[float], tester_name: str = "Тестируемый") -> str:
        if not results:
            raise ValueError("Нет данных для сохранения")
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        sheet_name = f"{tester_name}_{timestamp}"[:31]  # Excel ограничение 31 символ
        if sheet_name in self.wb.sheetnames:
            i = 1
            while f"{sheet_name}_{i}" in self.wb.sheetnames:
                i += 1
            sheet_name = f"{sheet_name}_{i}"

        ws = self.wb.create_sheet(title=sheet_name)
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Результаты тестирования реакции - {tester_name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Заголовки таблицы
        headers = ["№", "Дата", "Время", "Время реакции (мс)", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Данные
        avg_time = sum(results) / len(results)
        start_row = 4

        for i, reaction_time in enumerate(results, 1):
            now = datetime.datetime.now()

            if reaction_time > 0:
                status = "Успех"
                time_value = reaction_time
                status_color = "C6EFCE"  # Светло-зеленый
            elif reaction_time == -1:
                status = "Таймаут"
                time_value = 0
                status_color = "FFEB9C"  # Светло-желтый
            else:
                status = "Ошибка"
                time_value = 0
                status_color = "FFC7CE"  # Светло-красный

            ws.cell(row=start_row + i - 1, column=1, value=i)
            ws.cell(row=start_row + i - 1, column=2, value=now.date())
            ws.cell(row=start_row + i - 1, column=3, value=now.time())
            ws.cell(row=start_row + i - 1, column=4, value=time_value)
            ws.cell(row=start_row + i - 1, column=5, value=status)

            status_cell = ws.cell(row=start_row + i - 1, column=5)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

        # Статистика
        stats_row = start_row + len(results) + 2
        ws.merge_cells(f'A{stats_row}:E{stats_row}')
        stats_title = ws.cell(row=stats_row, column=1, value="СТАТИСТИКА")
        stats_title.font = Font(bold=True, size=12)
        successful = [r for r in results if r > 0]

        stats_data = [
            ["Общее количество тестов:", len(results)],
            ["Успешных тестов:", len(successful)],
            ["Процент успеха:", f"{(len(successful) / len(results)) * 100:.1f}%" if results else "0%"],
            ["Среднее время:", f"{avg_time:.0f} мс"],
            ["Лучшее время:", f"{min(successful):.0f} мс" if successful else "0 мс"],
            ["Худшее время:", f"{max(successful):.0f} мс" if successful else "0 мс"],
        ]

        for i, (label, value) in enumerate(stats_data, 1):
            ws.cell(row=stats_row + i, column=2, value=label).font = Font(bold=True)
            ws.cell(row=stats_row + i, column=3, value=value)

        self._auto_adjust_columns(ws)
        self.wb.save(self.filename)
        return self.filename

    def _auto_adjust_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def get_available_sheets(self) -> List[str]:
        return self.wb.sheetnames if self.wb else []

    def close(self):
        if self.wb:
            self.wb.close()


# Простая функция для обратной совместимости
def save_to_excel(results: List[float], tester_name: str = "Тестируемый",
                  filename: str = "test_results.xlsx") -> str:
    manager = ExcelManager(filename)
    return manager.save_results(results, tester_name)